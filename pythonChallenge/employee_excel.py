import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook

# Create a Tkinter window
root = tk.Tk()
root.title("Excel Spreadsheet Creation")

# Create an empty Excel workbook
workbook = Workbook()
employee_data = workbook.active
employee_data.title = "Employee Data"


# Function to populate and consolidate data
def populate_and_consolidate():
    # Define column headers
    headers = [
        "Employee ID",
        "First Name",
        "Last Name",
        "Department",
        "Total Amount",
        "Total Quantity",
    ]

    # Populate headers in the first row
    for col, header in enumerate(headers, 1):
        employee_data.cell(row=1, column=col, value=header)

    # Sample data for each product sale
    sample_data = [
        [342456, "Rodrigo", "Gregorio", "Electronics", 200, 3],
        # Add more sample data for other products as needed
    ]

    # Populate sample data
    for row, data in enumerate(sample_data, 2):
        for col, value in enumerate(data, 1):
            employee_data.cell(row=row, column=col, value=value)

    # Consolidate data based on Department
    consolidated_data = {}

    for row in employee_data.iter_rows(min_row=2, values_only=True):
        department = row[3]  # Department is in the 4th column (index 3)
        total_amount = row[4]  # Total Amount is in the 5th column (index 4)
        total_quantity = row[5]  # Total Quantity is in the 6th column (index 5)

        if department in consolidated_data:
            consolidated_data[department][0] += total_amount
            consolidated_data[department][1] += total_quantity
        else:
            consolidated_data[department] = [total_amount, total_quantity]

    # Write consolidated data
    for row, (department, (total_amount, total_quantity)) in enumerate(
        consolidated_data.items(), start=len(sample_data) + 2
    ):
        employee_data.cell(row=row, column=4, value=department)
        employee_data.cell(row=row, column=5, value=total_amount)
        employee_data.cell(row=row, column=6, value=total_quantity)

    # Save the spreadsheet as "employee_records.xlsx"
    save_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")]
    )
    if save_path:
        workbook.save(save_path)
        result_label.config(text=f"Spreadsheet saved as '{save_path}'")


# Create a button to populate and consolidate data
process_button = tk.Button(
    root, text="Populate and Consolidate", command=populate_and_consolidate
)
process_button.pack()

# Label to display the result or error message
result_label = tk.Label(root, text="")
result_label.pack()

root.mainloop()
